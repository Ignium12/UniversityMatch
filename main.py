from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from fuzzywuzzy import fuzz, process

if __name__ == '__main__':
    path = 'C:\\Users\\kevin\\PycharmProjects\\UniversityMatch\\Daten_Upwork_1.xlsm'

    wb = load_workbook(filename=path)
    ws = wb['Tabelle1']
    ws2 = wb['Tabelle4']

    universityDict = {}
    matches = []

    j = 1
    for cell in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        universityDict[ws2.cell(column=2, row=j).value] = ws2.cell(column=3, row=j).value.split(',')

        j += 1

    i = 1
    for cell in ws['DQ']:
        description = cell.value
        if description is not None:
            description = description.lower()
        matches = []
        setUniversityName = ""
        if len(setUniversityName) != 0:
            break

        for universityName, universityCityState in universityDict.items():
            if description is None:
                break
            if universityName.lower() in description:
                ws.cell(column=135, row=i).value = universityName
                setUniversityName = universityName
                print(universityName, description)
                break


            else:
                Token_Set_Ratio = fuzz.token_set_ratio(universityName.lower(), description)
                if Token_Set_Ratio >= 95:
                    ws.cell(column=135, row=i).value = universityName
                    setUniversityName = universityName
                    break

                if len(universityCityState) > 1 and setUniversityName == "":
                    if universityCityState[0].lower() in description:
                        matches.append(universityName)
                        continue
                    elif universityCityState[1].lower() in description:
                        matches.append(universityName)

        # print(description, " ",matches)
        if len(matches) == 1:
            ws.cell(column=135, row=i).value = matches[0]
        elif len(matches) > 1:
            highest = process.extractOne(description, matches, scorer=fuzz.token_set_ratio, score_cutoff=90 )
            if highest is not None:
                ws.cell(column=135, row=i).value = highest[0]

            # for entryMatch in matches:
            #     if description is not None:
            #         Token_Set_Ratio = fuzz.token_set_ratio(entryMatch, description)
            #         if Token_Set_Ratio >= 90:
            #             ws.cell(column=135, row=i).value = entryMatch
            #             print(description, "MATCHED", entryMatch)

        i += 1

    wb.save('Daten_Upwork_1.xlsx')
